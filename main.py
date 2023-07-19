import openpyxl
import datetime
from openpyxl import load_workbook

# Replace 'path/to/your/file.xlsx' with the actual path to your XLSX file
file_path = '/Users/hadi/PycharmProjects/JManTech500/slides_and_excel/230619 UpReach Bootcamp PowerBI and Python Challenge.xlsx'

# Load the Excel workbook
workbook = openpyxl.load_workbook(file_path)

# Select the first sheet (you can select other sheets by name if needed)
sheet = workbook["Customers"]
CustomerDict={}
# Iterate through the rows and columns and read the data
for row in sheet.iter_rows(min_row=1, values_only=True):
    if row[1] in CustomerDict:
        CustomerDict[row[1]].append(row[0])
    else:
        CustomerDict[row[1]]=[row[0]]
print(CustomerDict)

# Close the workbook after reading
workbook.close()
